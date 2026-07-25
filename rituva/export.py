"""Export a plan to .xlsx — Excel parity with the reference Eastern Diet Planner
(PRD §11.3 / goal G7). Uses openpyxl (optional dep). Sheets: Week · Grocery · Targets.

Returns raw bytes so the API can stream it as a download. Values come from the plan
payload (all DB-computed); this module formats, it does not compute nutrition.
"""
from __future__ import annotations

import io

from .grocery import aggregate


def plan_to_xlsx(plan: dict, member_name: str = "", people: int = 1) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    head = Font(bold=True, color="241300")
    fill = PatternFill("solid", fgColor="FFB020")
    wb = Workbook()

    # --- Week ---
    ws = wb.active
    ws.title = "Week"
    ws.append(["Date", "Season", "Breakfast", "Lunch", "Dinner", "Snack 1", "Snack 2", "kcal", "DQS"])
    for c in ws[1]:
        c.font, c.fill = head, fill
    slot_col = {"breakfast": 2, "lunch": 3, "dinner": 4, "snack1": 5, "snack2": 6}
    for d in plan.get("days", []):
        row = [d["date"], d.get("season", "")] + [""] * 5 + [
            round(d["totals"]["kcal"]), d["validation"]["dqs"]]
        for e in d["entries"]:
            idx = slot_col.get(e["slot"])
            if idx is not None:
                row[idx] = " + ".join(c["name"] for c in e["components"])
        ws.append(row)
    for i, w in enumerate([12, 10, 26, 32, 26, 18, 18, 7, 6], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # --- Grocery ---
    gs = wb.create_sheet("Grocery")
    gs.append(["Item", "Category", "Quantity", "Unit"])
    for c in gs[1]:
        c.font, c.fill = head, fill
    for cat in aggregate(plan, people=people)["categories"]:
        for it in cat["items"]:
            gs.append([it["item"], it["category"], it["quantity"], it["unit"]])
    for i, w in enumerate([28, 14, 10, 8], 1):
        gs.column_dimensions[chr(64 + i)].width = w

    # --- Targets ---
    tsh = wb.create_sheet("Targets")
    t = plan.get("targets", {})
    tsh.append(["Rituva plan for", member_name])
    tsh.append(["Metric", "Value"])
    for c in tsh[2]:
        c.font, c.fill = head, fill
    for k in ("kcal", "protein_g", "fat_g", "carb_g", "fibre_g",
              "sodium_mg_max", "added_sugar_g_max", "source"):
        tsh.append([k, t.get(k, "")])
    tsh.append(["citations", ", ".join(t.get("citations", []))])
    tsh.column_dimensions["A"].width = 20
    tsh.column_dimensions["B"].width = 44

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def plan_to_pdf(plan: dict, member_name: str = "", people: int = 1) -> bytes:
    """Render the plan + grocery as a printable PDF (fpdf2). Values are DB-computed;
    this only formats. Core fonts are latin-1, so text is sanitized to latin-1."""
    from fpdf import FPDF

    def s(x):
        return str(x).encode("latin-1", "replace").decode("latin-1")

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(20, 20, 16)
    pdf.cell(0, 10, s("Rituva - Weekly Plan"), ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(90, 90, 90)
    t = plan.get("targets", {})
    who = member_name or plan.get("member_id", "")
    pdf.cell(0, 6, s(f"{who}   target {t.get('kcal', '')} kcal - protein {t.get('protein_g', '')} g "
                     f"- {t.get('source', '')}"), ln=True)
    pdf.ln(2)

    order = ["breakfast", "snack1", "lunch", "snack2", "dinner"]
    for d in plan.get("days", []):
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(20, 20, 16)
        pdf.cell(0, 8, s(f"{d['date']}    {round(d['totals']['kcal'])} kcal - DQS {d['validation']['dqs']}"), ln=True)
        by = {e["slot"]: e for e in d["entries"]}
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(50, 50, 50)
        for slot in order:
            e = by.get(slot)
            if not e:
                continue
            names = " + ".join(c["name"] for c in e["components"])
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(0, 6, s(f"{slot.capitalize()}:  {names}  ({round(e['nutrients']['kcal'])} kcal)"))
        pdf.ln(1)

    pdf.add_page()
    g = aggregate(plan, people=people)
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(20, 20, 16)
    pdf.cell(0, 9, s(f"Grocery - {g['total_items']} items - {g['people']} people"), ln=True)
    for cat in g["categories"]:
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(30, 30, 30)
        pdf.cell(0, 7, s(cat["category"]), ln=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(50, 50, 50)
        for it in cat["items"]:
            pdf.cell(0, 5.5, s(f"   {it['item']}  -  {it['quantity']} {it['unit']}"), ln=True)

    pdf.ln(3)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.multi_cell(0, 4, s("All nutrient values are computed from the Knowledge DB (IFCT 2017 / DGI 2024). "
                           "Guideline-based general nutrition - not medical advice."))
    return bytes(pdf.output())
